from django.contrib.auth.models import Group, Permission
from django.core.files.uploadedfile import SimpleUploadedFile, InMemoryUploadedFile
from django.db.models.functions import Now
from django.test import TestCase, Client
from django.urls import reverse
from django.contrib.auth import get_user_model

from openpyxl import Workbook, load_workbook

import rm
from users.models import OrganizationalUnit
from .constants import ERROR_MSG_FILE_DEFINITION_ERROR, ERROR, OK, NEGOMETRIX, CONTRACTEN
from .interface_file_util import check_file_and_interface_type
from .models import Contract, InterfaceCall, System, DataSetType, InterfaceDefinition, DataPerOrgUnit, Mapping
from django.db.utils import IntegrityError

from .interface_file import check_file_is_excel_file, check_file_has_excel_extension, is_valid_header_row, \
    get_mandatory_field_positions, get_headers_from_sheet, get_fields_with_their_position, mandatory_fields_present, \
    get_org_unit
from .negometrix import register_contract, NegometrixInterfaceFile


def setUpUserWithInterfaceCallAndContract(self,
                                          superuser=True,
                                          group_name=None,
                                          username="John",
                                          name_in_negometrix="John"):
    """
    When called without parameters, you get a superuser
    """
    if superuser:
        self.user = _create_superuser()
    else:
        self.user = _create_user(group_name=group_name,
                                 username=username,
                                 name_in_negometrix=name_in_negometrix,
                                 )

    self.client.force_login(self.user)

    # Set up static data
    self.system_a = System.objects.create(name="SYSTEM_A")
    self.system = System.objects.create(name="Negometrix")
    self.dataset_type = DataSetType.objects.create(name="contracten")
    self.interface_definition = InterfaceDefinition.objects.create(system=self.system,
                                                                   dataset_type=self.dataset_type,
                                                                   interface_type=InterfaceDefinition.UPLOAD)
    self.org_unit = OrganizationalUnit.objects.create(name="MyTeam",
                                                      type=OrganizationalUnit.TEAM)

    # Set up process data
    self.interface_call = InterfaceCall.objects.create(date_time_creation=Now(),
                                                       status='TestStatus',
                                                       filename='Text.xls',
                                                       interface_definition=self.interface_definition)

    self.data_per_org_unit = DataPerOrgUnit.objects.create(interface_call=self.interface_call,
                                                           org_unit=self.org_unit)

    self.contract_1 = Contract.objects.create(contract_nr='NL-123',
                                              seq_nr=0,
                                              description='Test Contract 1',
                                              contract_owner='T. Ester',
                                              contract_name='Test contract naam',
                                              data_per_org_unit=self.data_per_org_unit)


def _create_superuser(username="john", password="doe", **kwargs):
    user = get_user_model().objects.create(username=username,
                                           is_superuser=True,
                                           is_active=True,
                                           **kwargs
                                           )
    if password:
        user.set_password(password)
    else:
        user.set_unusable_password()
    user.save()
    return user


def _create_user(username="john",
                 password="doe",
                 group_name=None,
                 name_in_negometrix="J. Doe",
                 **kwargs):
    user = get_user_model().objects.create(username=username,
                                           is_active=True,
                                           **kwargs
                                           )

    user.name_in_negometrix = name_in_negometrix
    if password:
        user.set_password(password)
    else:
        user.set_unusable_password()

    if group_name:
        try:
            group = Group.objects.get(name=group_name)
        except Group.DoesNotExist:
            message = f"Group {group_name} can not be found"
            raise Exception(message)
        user.groups.add(group)

    user.save()

    return user


class DataModelTest(TestCase):

    def setUp(self):
        setUpUserWithInterfaceCallAndContract(self)

    def testContractsOfInterfaceCall_one_contract(self):
        contracts = self.interface_call.contracts()
        self.assertEqual(len(contracts), 1)

    def testContractsOfInterfaceCall_two_contracts(self):
        self.contract_2 = Contract.objects.create(contract_nr='NL-123',
                                                  seq_nr=0,
                                                  description='Test Contract 2',
                                                  contract_owner='T. Ester',
                                                  contract_name='Test contract naam',
                                                  data_per_org_unit=self.data_per_org_unit)

        contracts = self.interface_call.contracts()
        self.assertEqual(len(contracts), 2)

class ContractTest(TestCase):

    def setUp(self):
        setUpUserWithInterfaceCallAndContract(self)

    def test_homepage(self):
        c = self.client
        response = c.get("/")
        self.assertEqual(response.status_code, 200)

    def test_contract(self):
        expected = 'NL-123: Test contract naam'
        self.assertEqual(self.contract_1.__str__(), expected)

    def test_one_interface_call_on_page(self):
        response = self.client.get('/interfacecalls')
        self.assertContains(response, 'TestStatus')

    def test_one_contract_on_interface_call_page(self):
        response = self.client.get(f'/interfacecall/{self.interface_call.pk}/')
        self.assertContains(response, 'NL-123')
        self.assertContains(response, 'Test Contract')
        self.assertContains(response, 'T. Ester')

    def test_two_contract_on_interface_call_page(self):
        Contract.objects.create(contract_nr='NL-345',
                                seq_nr=1,
                                description='Test Contract 2',
                                contract_owner='T. Ester',
                                data_per_org_unit=self.data_per_org_unit)
        response = self.client.get(f'/interfacecall/{self.interface_call.pk}/')
        self.assertContains(response, 'NL-123')
        self.assertContains(response, 'Test Contract')
        self.assertContains(response, 'T. Ester', count=2)

        self.assertContains(response, 'NL-345')
        self.assertContains(response, 'Test Contract 2')

    def test_create_contract_without_parent(self):
        try:
            Contract.objects.create(seq_nr=0, contract_nr="NL-123", data_per_org_unit=self.data_per_org_unit)
        except IntegrityError as exception:
            expected = "null value in column \"interface_call_id\" " \
                       "violates not-null constraint"
            self.assertTrue(expected in exception.__str__())

class ExcelTests(TestCase):

    # Excel file extension

    def setUp(self):
        setUpUserWithInterfaceCallAndContract(self)

    def test_not_an_excel_file_extension(self):
        filename = 'test.txt'
        try:
            check_file_has_excel_extension(filename)
        except Exception as ex:
            self.assertTrue("Bestand heeft geen excel extensie" in ex.__str__())

    def test_an_excel_file_extension_xls(self):
        filename = 'test.xls'
        self.assertTrue(check_file_has_excel_extension(filename))

    def test_an_excel_file_extension_XLS(self):
        filename = 'test.XLS'
        self.assertTrue(check_file_has_excel_extension(filename))

    def test_an_excel_file_extension_xlsx(self):
        filename = 'test.xlsx'
        self.assertTrue(check_file_has_excel_extension(filename))

    def test_an_excel_file_extension_XLSX(self):
        filename = 'test.XLSX'
        self.assertTrue(check_file_has_excel_extension(filename))

    # Test on really being an Excel file, so can it be read as Workbook?

    def test_empty_file_with_xls_extension(self):
        file = open("rm/test/resources/EmptyFileWithXLSExtension.xls", "rb")
        print(f'type of file: {type(file)}')
        try:
            is_excel = check_file_is_excel_file(file)
        except Exception as ex:
            self.assertTrue("Het openen van dit bestand als excel bestand"
                            " geeft een foutmelding" in ex.__str__())

    def test_not_an_excelfile(self):
        file = open("rm/test/resources/aPdfWithExcelExtension.xls", "rb")
        print(f'type of file: {type(file)}')
        try:
            is_excel = check_file_is_excel_file(file)
        except Exception as ex:
            self.assertTrue("Het openen van dit bestand als excel bestand"
                            " geeft een foutmelding" in ex.__str__())

    def test_an_excelfile_xlsx(self):
        file = open("rm/test/resources/a_valid_excel_file.xlsx", "rb")
        print(f'type of file: {type(file)}')
        is_excel = check_file_is_excel_file(file)
        self.assertTrue(is_excel)

    # Test valid headers

    def test_valid_header_row_exact(self):
        self.assertTrue(is_valid_header_row(found_headers=('x', 'y', 'z'),
                                            mandatory_headers=('x', 'z')))

    def test_valid_header_row_case_sensitive(self):
        self.assertTrue(is_valid_header_row(found_headers=('X', 'y', 'z'),
                                            mandatory_headers=('x', 'z')))

    def test_invalid_header_row(self):
        self.assertFalse(is_valid_header_row(found_headers=('a', 'b', 'z'),
                                             mandatory_headers=('x', 'z')))

    # Test handle uploaded excel file

    def test_check_excel_file_invalid_headers(self):
        try:
            interfaceCall = InterfaceCall.objects.create(date_time_creation=Now(),
                                                         status='TestStatus',
                                                         filename='valid_excel_without_headers.xlsx',
                                                         interface_definition=self.interface_definition)
            file = "rm/test/resources/valid_excel_without_headers.xlsx"
            check_file_and_interface_type(file)
        except Exception as ex:
            self.assertTrue("File can not be recognized" in ex.__str__(),
                            f"We got another exception than expected, received: {ex.__str__()}")
        else:
            self.assertTrue(False, "No Exception, while expected because file has no headers")


    #  Test generic database functions

    def test_get_field_positions(self):

        found_headers = ['Header 1', "just a field", 'Header x', 'and another']
        defined_headers = dict(
            header_01="Header 1",
            zomaarwat="something",
            header_x="Header x",
        )
        field_positions = get_fields_with_their_position(found_headers,
                                                         defined_headers)

        expected_field_opsitions = dict(
            header_01=0,
            header_x=2,
        )
        self.assertEqual(len(field_positions), 2)
        self.assertEqual(field_positions, expected_field_opsitions)

    def test_get_mandatory_field_positions_valid_situation(self):
        mandatory_fields = ("x", "z")
        field_positions = {"x": 1, "y": 2, "z": 4}
        positions = get_mandatory_field_positions(mandatory_fields,
                                                  field_positions)
        expected = [1, 4]
        self.assertEqual(positions, expected)

    def test_get_mandatory_field_postions_invalid_situation(self):
        mandatory_fields = ("x", "z")
        field_positions = {"x": 1, "y": 2, "b": 3}
        try:
            get_mandatory_field_positions(mandatory_fields,
                                          field_positions)
        except Exception as ex:
            self.assertEqual(ex.__str__(), ERROR_MSG_FILE_DEFINITION_ERROR)

    def test_mandatory_fields_present(self):
        mandatory_field_positions = [2, 3]
        row_values = ["nul", "een", "twee", "drie", "vier"]
        self.assertTrue(mandatory_fields_present(mandatory_field_positions,
                                                 row_values))

    def test_mandatory_fields_not_present_1_pos(self):
        mandatory_field_positions = [3]
        row_values = ["nul", "een", "twee", None, "vier"]
        self.assertFalse(mandatory_fields_present(mandatory_field_positions,
                                                  row_values))

    def test_mandatory_fields_not_present_1_ok_1_not(self):
        mandatory_field_positions = [2, 3]
        row_values = ["nul", "een", "twee", None, "vier"]
        self.assertFalse(mandatory_fields_present(mandatory_field_positions,
                                                  row_values))

    def test_get_headers(self):
        file = open("rm/test/resources/test_get_headers.xlsx", "rb")
        workbook: Workbook = load_workbook(file)
        sheet = workbook.active
        headers = get_headers_from_sheet(sheet)
        expected = ("HEADER 1", "HEADER 2", None, "HEADER 4")
        self.assertEqual(headers, expected)

    def test_register_contract(self):
        row_nr = 4
        file = open("rm/test/resources/test_register_contract.xlsx", "rb")
        workbook: Workbook = load_workbook(file)
        sheet = workbook.active
        count = 1
        values_row_4 = []
        for row_values in sheet.iter_rows(min_row=1,
                                          min_col=1,
                                          values_only=True):
            if count == 4:
                values_row_4 = row_values
                break
            count += 1

        interfaceCall = InterfaceCall.objects.create(date_time_creation=Now(),
                                                     status='TestStatus',
                                                     filename='test_register_contract.xlsx',
                                                     interface_definition=self.interface_definition)

        fields_with_position = dict(database_nr=0,
                                    contract_nr=1,
                                    contract_status=2,
                                    description=3,
                                    )

        mandatory_field_positions = (1, 2)
        mandatory_fields = ('a_field', 'another_field')

        status, msg = register_contract(row_nr, values_row_4, interfaceCall, fields_with_position,
                                        mandatory_field_positions)

        expected_status = ERROR

        self.assertEqual(status, expected_status)

    def test_get_org_unit_found_by_system(self):
        mapping_name = "my_map"
        mapping = Mapping.objects.create(name=mapping_name,
                                         system=self.system_a,
                                         org_unit=self.org_unit)
        found_org_unit = get_org_unit(self.system_a, mapping_name)
        self.assertEqual(self.org_unit.name, found_org_unit.name)

    def test_get_org_unit_found_by_system_name(self):
        mapping_name = "my_map"
        mapping = Mapping.objects.create(name=mapping_name,
                                         system=self.system_a,
                                         org_unit=self.org_unit)
        found_org_unit = get_org_unit(self.system_a.name, mapping_name)
        self.assertEqual(self.org_unit.name, found_org_unit.name)

    def test_get_org_unit_unknown_system_name(self):
        mapping_name = "my_map"
        mapping = Mapping.objects.create(name=mapping_name,
                                         system=self.system_a,
                                         org_unit=self.org_unit)
        found_org_unit = get_org_unit("SYSTEM UNKNOWN", mapping_name)
        self.assertIsNone(found_org_unit)

    def test_get_org_unit_unknown_mapping(self):
        not_existing_mapping_name = "Not existing"
        mapping_name = "my_map"
        mapping = Mapping.objects.create(name=mapping_name,
                                         system=self.system_a,
                                         org_unit=self.org_unit)
        found_org_unit = get_org_unit(self.system_a, not_existing_mapping_name)
        self.assertIsNone(found_org_unit)




# class FileUploadTests(TestCase):
#
#     def setUp(self):
#         setUpUserWithInterfaceCallAndContract(self, superuser=True)
#
#     def test_upload_empty_file_with_xls(self):
#
#         nr_int_calls_before = len(InterfaceCall.objects.all())
#
#         file = SimpleUploadedFile(name="rm/test/resources/EmptyFileWithXLSExtension.xls",
#                                   content=b"file_content",
#                                   content_type="excel/xls")
#         response = self.client.post(reverse('upload'), {'file': file})
#         nr_int_calls_after = len(InterfaceCall.objects.all())
#         self.assertEqual(nr_int_calls_after, nr_int_calls_before + 1)
#
#         interface_call: InterfaceCall = InterfaceCall.objects.last()
#         self.assertEqual(interface_call.filename, "EmptyFileWithXLSExtension.xls")
#         self.assertTrue(interface_call.message.__contains__('Het openen van dit bestand als excel bestand '
#                                                             'geeft een foutmelding: File is not a zip file'))
#
#         self.assertContains(response, 'Het openen van dit bestand als excel bestand geeft een '
#                                       'foutmelding: File is not a zip file')
#
#
#     def test_upload_a_valid_excel_file(self):
#
#         nr_int_calls_before = len(InterfaceCall.objects.all())
#
#         file = SimpleUploadedFile(name="rm/test/resources/a_valid_excel_file.xlsx",
#                                   content=b"file_content",
#                                   content_type="excel/xls")
#         response = self.client.post(reverse('upload'), {'file': file})
#         nr_int_calls_after = len(InterfaceCall.objects.all())
#         self.assertEqual(nr_int_calls_after, nr_int_calls_before + 1)
#
#         interface_call: InterfaceCall = InterfaceCall.objects.last()
#         self.assertEqual(interface_call.filename, "a_valid_excel_file.xlsx")
#         self.assertEqual(interface_call.status, "OK", msg=interface_call.message)
#         self.assertIsNone(interface_call.message)
#
#
    # def test_upload_Overzicht_Tech_Cluster_Backend_2_12_xlsx(self):
    #
    #     nr_int_calls_before = len(InterfaceCall.objects.all())
    #
    #     with open('rm/test/resources/Overzicht_Tech_Cluster_Backend 2-12.xlsx') as upload_file:
    #         response = self.client.post(reverse('upload'),
    #                                     {'name': 'rm/test/resources/Overzicht_Tech_Cluster_Backend 2-12.xlsx',
    #                                      'content': b"file_content",
    #                                      'attachment': upload_file})
    #
    #     # file = SimpleUploadedFile(name="rm/test/resources/Overzicht_Tech_Cluster_Backend 2-12.xlsx",
    #     #                           content=b"file_content",
    #     #                           content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    #     #
    #     # response = self.client.post(reverse('upload'), {'file': file})
    #     nr_int_calls_after = len(InterfaceCall.objects.all())
    #     self.assertEqual(nr_int_calls_after, nr_int_calls_before + 1)
    #
    #     interface_call: InterfaceCall = InterfaceCall.objects.last()
    #     self.assertEqual(interface_call.filename, "Overzicht_Tech_Cluster_Backend 2-12.xlsx")
    #     self.assertEqual(interface_call.status, "OK", msg=interface_call.message)
    #     self.assertIsNone(interface_call.message)


class LoginRequiredTests(TestCase):
    """
    All pages require login, except the login page
    """

    def test_login_required_home_page(self):
        response = self.client.get(reverse("home"))
        self.assertRedirects(response, reverse('account_login') + "?next=/")

    def test_login_required_upload_page(self):
        response = self.client.get(reverse("upload"))
        self.assertRedirects(response, reverse('account_login') + "?next=/upload/")


class RoleBasedAuthorizationSuperuserTests(TestCase):

    def setUp(self):
        setUpUserWithInterfaceCallAndContract(self, superuser=True)

    def test_superuser_sees_upload_button(self):
        response = self.client.get(reverse("interface_call_list"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Upload file')

    def test_superuser_can_access_upload_form(self):
        response = self.client.get(reverse("upload"))
        self.assertEqual(response.status_code, 200)

    def test_superuser_sees_contracts_of_interfaceCall(self):
        response = self.client.get(f'/interfacecall/{self.interface_call.pk}/')
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'NL-123')
        self.assertContains(response, 'Test Contract')
        self.assertContains(response, 'T. Ester', count=1)


def print_permissions_and_groups():
    all_permissions = Permission.objects.all()
    print("--------------------------------------")
    print("           PERMISSIONS")
    print("--------------------------------------")
    for permission in all_permissions:
        print(f"Found permission: {permission.codename}")
    print("--------------------------------------")

    all_groups = Group.objects.all()
    print("--------------------------------------")
    print("           GROUPS")
    print("--------------------------------------")
    for group in all_groups:
        print(f"Found group: {group.name}")
        permissions = group.permissions.all()
        if len(permissions) == 0:
            print("- has no permissions")
        for permission in permissions:
            print(f"- has permission: {permission.codename}")
    print("--------------------------------------")


class RoleBasedAuthorizationClusterLeadTests(TestCase):

    def setUp(self):
        setUpUserWithInterfaceCallAndContract(self, superuser=False, group_name="Cluster Lead")
        # print_permissions_and_groups()

    def test_cluster_lead_sees_no_upload_button(self):
        response = self.client.get(reverse("interface_call_list"))
        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, 'Upload file</a>')

    def test_cluster_lead_can_not_access_upload_form(self):
        response = self.client.get(reverse("upload"))
        self.assertEqual(response.status_code, 403)

    def test_cluster_lead_has_view_contract_permissions(self):
        permissions = self.user.user_permissions.all()
        self.assertTrue(self.user.has_perm('rm.view_contract'))

    def test_cluster_lead_sees_contracts_of_interfaceCall(self):
        response = self.client.get(f'/interfacecall/{self.interface_call.pk}/')
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'NL-123')
        self.assertContains(response, 'Test Contract')
        self.assertContains(response, 'T. Ester', count=1)


class RoleBasedAuthorizationBuyerTests(TestCase):

    def setUp(self):
        setUpUserWithInterfaceCallAndContract(self, superuser=False, group_name="Buyer")

    def test_buyer_sees_upload_button(self):
        response = self.client.get(reverse("interface_call_list"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Upload file')

    def test_buyer_can_access_upload_form(self):
        response = self.client.get(reverse("upload"))
        self.assertEqual(response.status_code, 200)

    def test_buyer_has_the_right_permissions(self):
        permissions = self.user.user_permissions.all()
        self.assertTrue(self.user.has_perm('rm.view_contract'))
        self.assertTrue(self.user.has_perm('rm.upload_contract_file'))
        self.assertTrue(self.user.has_perm('rm.call_contract_interface'))

    def test_buyer_sees_contracts_of_interfaceCall(self):
        response = self.client.get(f'/interfacecall/{self.interface_call.pk}/')
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'NL-123')
        self.assertContains(response, 'Test Contract 1')
        self.assertContains(response, 'T. Ester', count=1)
