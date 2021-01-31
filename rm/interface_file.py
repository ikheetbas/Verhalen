import logging
import pathlib
from abc import ABC, abstractmethod
from typing import Tuple, Dict, List

from django.core.files.uploadedfile import UploadedFile
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

import rm
from rm.constants import ERROR_MSG_FILE_DEFINITION_ERROR, TOTAL_ROWS_RECEIVED, TOTAL_DATA_ROWS_RECEIVED, \
    RowStatus, FileStatus
from rm.models import InterfaceCall, RawData, System, Mapping

logger = logging.getLogger(__name__)


def check_file_has_excel_extension(filename: str) -> bool:
    """
    Returns True if file has an Excel extension, if not, an exception is thrown.
    """
    extension = pathlib.Path(filename).suffix.upper()
    if extension in ['.XLS', '.XLSX']:
        return True
    else:
        raise Exception(f"Bestand heeft geen excel extensie maar: '{extension}'")


def check_file_is_excel_file(file: UploadedFile) -> bool:
    """
    Returns True if the file is a valid Excel File. If not, an exception is thrown.
    """
    try:
        load_workbook(filename=file)
        return True
    except Exception as ex:
        raise Exception(f"Het openen van dit bestand als excel bestand"
                        f" geeft een foutmelding: {ex.__str__()}")


def get_headers_from_sheet(sheet: Worksheet) -> Tuple[str]:
    """
    Returns a Tuple with the cell content of the 1st row, empty cells have a
    None value. Example: ("Contract nr.", "Contract Status", None, "Eigenaar")
    """
    headers = []
    for value in sheet.iter_rows(min_row=1,
                                 max_row=1,
                                 values_only=True):
        headers.append(value)
    return headers[0]


def get_headers_from_file(file) -> Tuple[str]:
    workbook = load_workbook(filename=file)
    sheet = workbook.active
    found_headers = get_headers_from_sheet(sheet)
    return found_headers


def to_upper_none_proof(x):
    if x and hasattr(x, "upper"):
        return x.upper()
    return x


def is_valid_header_row(found_headers: Tuple[str],
                        mandatory_headers: Tuple[str]) -> bool:
    """
    Checks if the found headers contain all mandatory headers.
    """
    headers_upper = [to_upper_none_proof(x) for x in found_headers]
    required_headers_upper = [x.upper() for x in mandatory_headers]
    valid = all(item in headers_upper for item in required_headers_upper)
    return valid


def get_mandatory_field_positions(mandatory_fields: Tuple[str, ...],
                                  field_positions: Dict[str, int]) -> List[int]:
    """
    Returns the position in the row that must have a value.
    """
    positions = []

    for fieldname in mandatory_fields:
        if fieldname in field_positions:
            positions.append(field_positions[fieldname])

    if len(mandatory_fields) > len(positions):
        logger.error(ERROR_MSG_FILE_DEFINITION_ERROR +
                     mandatory_fields.__str__() + " " +
                     field_positions.__str__())
        raise Exception(ERROR_MSG_FILE_DEFINITION_ERROR)

    return positions


def replace_none_with_blank_and_make_50_long(row_value):
    values = []
    for cell_value in row_value:
        values.append("" if not cell_value else cell_value)
    while len(values) < 50:
        values.append("")
    return values


def row_is_empty(row_values: Tuple[str]) -> bool:
    return row_values.count(None) == len(row_values)


def get_fields_with_their_position(found_headers, defined_headers):
    """
    Returns the positions, as defined in the defined_headers, of the available headers
    as found in the file. First position has index: 0
    """
    field_positions = dict()

    for fieldname in defined_headers:
        defined_header = defined_headers[fieldname]
        if defined_header in found_headers:
            position = found_headers.index(defined_header)
            field_positions[fieldname] = position

    return field_positions


def mandatory_fields_present(mandatory_field_positions: Tuple[int, ...],
                             row_values: Tuple[str, ...]) -> bool:
    """
    Checks in the mandatory fields in the row have a value. If a mandatory field contains
    a None or "" a False is returned.
    """
    for position in mandatory_field_positions:
        value_to_be_checked = row_values[position]
        if not value_to_be_checked:
            return False
        if row_values[position] == "":
            return False
    return True


def register_in_raw_data(row_nr: int,
                         row_value: Tuple[str],
                         interfaceCall: InterfaceCall) -> RawData:
    values = replace_none_with_blank_and_make_50_long(row_value)
    raw_data = RawData.objects.create(interface_call=interfaceCall,
                                      seq_nr=row_nr,
                                      status='NEW',
                                      field_01=values[0],
                                      field_02=values[1],
                                      field_03=values[2],
                                      field_04=values[3],
                                      field_05=values[4],
                                      field_06=values[5],
                                      field_07=values[6],
                                      field_08=values[7],
                                      field_09=values[8],
                                      field_10=values[9],
                                      field_11=values[10],
                                      field_12=values[11],
                                      field_13=values[12],
                                      field_14=values[13],
                                      field_15=values[14],
                                      field_16=values[15],
                                      field_17=values[16],
                                      field_18=values[17],
                                      field_19=values[18],
                                      field_20=values[19],
                                      field_21=values[20],
                                      field_22=values[21],
                                      field_23=values[22],
                                      field_24=values[23],
                                      field_25=values[24],
                                      field_26=values[25],
                                      field_27=values[26],
                                      field_28=values[27],
                                      field_29=values[28],
                                      field_30=values[29],
                                      field_31=values[30],
                                      field_32=values[31],
                                      field_33=values[32],
                                      field_34=values[33],
                                      field_35=values[34],
                                      field_36=values[35],
                                      field_37=values[36],
                                      field_38=values[37],
                                      field_39=values[38],
                                      field_40=values[39],
                                      field_41=values[40],
                                      field_42=values[41],
                                      field_43=values[42],
                                      field_44=values[43],
                                      field_45=values[44],
                                      field_46=values[45],
                                      field_47=values[46],
                                      field_48=values[47],
                                      field_49=values[48],
                                      field_50=values[49],
                                      )
    return raw_data


def get_org_unit(system, mapping_name):
    """
    Returns the OrganizationalUnit that can be found for this System with
    """
    if not isinstance(system, System):
        try:
            system = System.objects.get(name=system)
        except rm.models.System.DoesNotExist:
            return None

    try:
        mapping = Mapping.objects.get(system=system, name=mapping_name)
    except rm.models.Mapping.DoesNotExist:
        return None

    org_unit = mapping.org_unit

    return org_unit


class RowStatistics(object):

    def __init__(self):
        self.number_of_rows = {TOTAL_ROWS_RECEIVED: 0,
                               TOTAL_DATA_ROWS_RECEIVED: 0,
                               RowStatus.EMPTY_ROW: 0,
                               RowStatus.HEADER_ROW: 0,
                               RowStatus.DATA_OK: 0,
                               RowStatus.DATA_ERROR: 0,
                               RowStatus.DATA_WARNING: 0,
                               RowStatus.DATA_IGNORED: 0}

    def add_row_with_status(self, row_status):

        if isinstance(row_status,str):
            row_status = RowStatus[row_status]

        """
        Keep the statistics for all the row-statuses
        """
        self.number_of_rows[TOTAL_ROWS_RECEIVED] += 1
        self.number_of_rows[row_status] += 1
        if row_status.is_data_row():
            self.number_of_rows[TOTAL_DATA_ROWS_RECEIVED] += 1

    def get_total_rows_received(self):
        return self.number_of_rows[TOTAL_ROWS_RECEIVED]

    def get_total_data_rows_received(self):
        return self.number_of_rows[TOTAL_DATA_ROWS_RECEIVED]

    def get_total_empty_rows(self):
        return self.number_of_rows[RowStatus.EMPTY_ROW]

    def get_total_header_rows(self):
        return self.number_of_rows[RowStatus.HEADER_ROW]

    def get_total_data_ok_rows(self):
        return self.number_of_rows[RowStatus.DATA_OK]

    def get_total_data_error_rows(self):
        return self.number_of_rows[RowStatus.DATA_ERROR]

    def get_total_data_warning_rows(self):
        return self.number_of_rows[RowStatus.DATA_WARNING]

    def get_total_data_ignored_rows(self):
        return self.number_of_rows[RowStatus.DATA_IGNORED]

class ExcelInterfaceFile(ABC):
    """
    Base class containing all non-specific logic for uploading an Excel file
    with data.
    """
    interfaceCall: InterfaceCall = None

    row_statistics = RowStatistics()

    def __init__(self,
                 file):
        self.file = file

    def process(self, interfaceCall: InterfaceCall):
        self.interfaceCall = interfaceCall

        try:
            found_headers = get_headers_from_file(self.file)

            fields_with_their_position = self.get_fields_with_their_position(found_headers)

            self.handle_file(self.file,
                             self.interfaceCall,
                             fields_with_their_position)

            self.interfaceCall.status = FileStatus.OK.name

        except Exception as ex:
            self.interfaceCall.status = FileStatus.ERROR.name
            self.interfaceCall.message = f'Reason: {ex.__str__()}'


        self.interfaceCall.number_of_rows_received = self.row_statistics.get_total_rows_received()
        self.interfaceCall.number_of_empty_rows = self.row_statistics.get_total_empty_rows()
        self.interfaceCall.number_of_header_rows = self.row_statistics.get_total_header_rows()
        self.interfaceCall.number_of_data_rows_ok = self.row_statistics.get_total_data_ok_rows()
        self.interfaceCall.number_of_data_rows_ignored = self.row_statistics.get_total_data_ignored_rows()
        self.interfaceCall.number_of_data_rows_error = self.row_statistics.get_total_data_error_rows()
        self.interfaceCall.number_of_data_rows_received = self.row_statistics.get_total_data_rows_received()
        self.interfaceCall.number_of_data_rows_warning = self.row_statistics.get_total_data_warning_rows()

        self.interfaceCall.save()

        if not self.interfaceCall.status == FileStatus.OK.name:
            raise Exception(self.interfaceCall.message)

    @abstractmethod
    def get_interface_definition(self):
        raise Exception("Must be overridden, programming error")

    @abstractmethod
    def get_fields_with_their_position(self, found_headers: Tuple[str]) \
            -> Dict[str, int]:
        """
        With the headers out of the file, determine (with the help of the
        definitions of fieldnames and their headers) on which position on the row
        we can find which field. For example:
        - found_headers: "Contract nr." and "Contract owner"
        - in the file definition:
            contract_nr = "Contract nr."
            contract_owner = "Contract owner"
        - we return: (contract_nr = 0, contract_owner = 1)
        """
        raise Exception("Must be overridden, programming error")

    def handle_file(self,
                    file,
                    interface_call: InterfaceCall,
                    field_positions: Dict[str, int]):
        """
        Generic handling of the file.
        """

        self.interfaceCall = interface_call

        mandatory_field_positions: Tuple[int, ...] = \
            get_mandatory_field_positions(self.get_mandatory_fields(),
                                          field_positions)

        workbook = load_workbook(filename=file)
        sheet = workbook.active

        row_nr = 1
        for row_values in sheet.iter_rows(min_row=1,
                                          min_col=1,
                                          values_only=True):
            # TODO test on number of complete empty rows and STOP!
            row_status = self.handle_row(row_nr,
                                         row_values,
                                         field_positions,
                                         mandatory_field_positions)
            row_nr += 1
            self.row_statistics.add_row_with_status(row_status)

    def handle_row(self,
                   row_nr: int,
                   row_values: Tuple[str, ...],
                   field_positions: Dict[str, int],
                   mandatory_field_positions: Tuple[int, ...]) -> str:
        logger.debug(f"register RawData {row_nr} - {row_values}")

        raw_data = register_in_raw_data(row_nr,
                                        row_values,
                                        self.interfaceCall)
        try:
            status, message = self.register_business_data(row_nr,
                                                          row_values,
                                                          self.interfaceCall,
                                                          field_positions,
                                                          mandatory_field_positions)
        except Exception as ex:
            raw_data.status = RowStatus.DATA_ERROR.name
            raw_data.message = str(ex)
        else:
            raw_data.status = status.name
            raw_data.message = message
        logger.debug(f"Result register Contact {row_nr} : {raw_data.status} : {raw_data.message}")
        raw_data.save()
        return raw_data.status

    @abstractmethod
    def register_business_data(self,
                               row_nr: int,
                               row_values: Tuple[str, ...],
                               interface_call: InterfaceCall,
                               fields_with_position: Dict[str, int],
                               mandatory_field_positions: Tuple[int, ...]) -> Tuple[RowStatus, str]:
        """
        Handling of the row with data, creating the Business Data, to be implemented for each
        file type specific.
        """
        raise Exception("Must be overridden, programming error")

    @abstractmethod
    def get_mandatory_fields(self):
        """
        Each subclass has its own mandatory fields.
        """
        raise Exception("Must be overridden, programming error")


def fill_fields_in_record_from_row_values(record, fields_with_position, row_values):
    """
    Lookup all fields by position in row_values and store them in the record.
    """
    for field in fields_with_position:
        position = fields_with_position[field]
        value = row_values[position]
        setattr(record, field, value)