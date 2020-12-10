import logging
import pathlib
from typing import Tuple, Dict

from django.core.files.uploadedfile import UploadedFile
from django.db.models.functions import Now
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from rm.constants import ERROR, OK, NEW, ERROR_MSG_FILE_DEFINITION_ERROR, INTERFACE_TYPE_FILE
from rm.models import InterfaceCall, ReceivedData
from rm.negometrix import mandatory_headers, defined_headers, register_contract, mandatory_fields

logger = logging.getLogger(__name__)


def file_has_excel_extension(filename: str) -> [bool, str]:
    extension = pathlib.Path(filename).suffix.upper()
    if not extension in ['.XLS', '.XLSX']:
        return [False,
                f"Bestand heeft geen excel extensie maar: '{extension}'"]
    else:
        return [True, "OK"]


def file_is_excel_file(file: UploadedFile) -> [bool, str]:
    try:
        load_workbook(filename=file)
    except Exception as ex:
        return [False, f"Het openen van dit bestand als excel bestand"
                       f" geeft een foutmelding: {ex.__str__()}"]
    return [True, "OK"]


def get_headers(sheet: Worksheet) -> Tuple[str]:
    """
    Returns a Tuple with the cell content of the 1st row, empty cells have a
    None value. Example: ("Contract nr.", "Contract Status", None, "Eigenaar")
    """
    headers = []
    for value in sheet.iter_rows(min_row=1,
                                 max_row=1,
                                 values_only=True):
        headers.append(value)
    return headers[0]


def to_upper_none_proof(x):
    if x and hasattr(x, "upper"):
        return x.upper()
    return x


def is_valid_header_row(headers: Tuple[str], required_headers: Tuple[str]) -> bool:
    headers_upper = [to_upper_none_proof(x) for x in headers]
    required_headers_upper = [x.upper() for x in required_headers]
    valid = all(item in headers_upper for item in required_headers_upper)
    return valid


def get_field_positions(available_headers: Tuple[str],
                        defined_headers: Dict[str, int]) \
        -> Dict[str, int]:
    """
    Returns the positions, as defined in the defined_headers, of the available headers
    as found in the file. First position has index: 0
    """
    field_positions = dict()

    for fieldname in defined_headers:
        defined_header = defined_headers[fieldname]
        if defined_header in available_headers:
            position = available_headers.index(defined_header)
            field_positions[fieldname] = position

    return field_positions


def handle_uploaded_excel_file(excelfile, interfaceCall: InterfaceCall):
    workbook = load_workbook(filename=excelfile)
    sheet = workbook.active
    available_headers = get_headers(sheet)
    if not is_valid_header_row(available_headers, mandatory_headers):
        if hasattr(excelfile, "name"):  # when testing, name attribute is NA
            name = excelfile.name
        else:
            name = excelfile
        raise Exception(f"File '{name}' has no (valid) header row, "
                        f"missing one of {mandatory_headers}")

    field_positions = get_field_positions(available_headers, defined_headers)

    handle_negometrix_file(sheet, interfaceCall, field_positions)


def get_mandatory_field_positions(mandatory_fields: Tuple[str],
                                  field_positions: Dict[str, int]) -> Tuple[int]:
    positions = []

    for fieldname in mandatory_fields:
        if fieldname in field_positions:
            positions.append(field_positions[fieldname])

    if len(mandatory_fields) > len(positions):
        logger.error(ERROR_MSG_FILE_DEFINITION_ERROR +
                     mandatory_fields.__str__() + " " +
                     field_positions.__str__())
        raise Exception(ERROR_MSG_FILE_DEFINITION_ERROR)

    return positions


def handle_negometrix_file(sheet: Worksheet,
                           interfaceCall: InterfaceCall,
                           field_positions: Dict[str, int]):
    mandatory_field_positions: Tuple[int] = \
        get_mandatory_field_positions(mandatory_fields,
                                      field_positions)

    row_nr = 1
    for row_values in sheet.iter_rows(min_row=1,
                                      min_col=1,
                                      values_only=True):
        # TODO test on complete empty row and STOP!
        handle_negometrix_row(row_nr,
                              row_values,
                              interfaceCall,
                              field_positions,
                              mandatory_field_positions)
        row_nr += 1


def replace_none_with_blank_and_make_50_long(row_value):
    values = []
    for cell_value in row_value:
        values.append("" if not cell_value else cell_value)
    while len(values) < 50:
        values.append("")
    return values


def register_in_recevied_data(rownr: int,
                              row_value: Tuple[str],
                              interfaceCall: InterfaceCall) -> ReceivedData:
    values = replace_none_with_blank_and_make_50_long(row_value)
    receivedData = ReceivedData.objects.create(interface_call=interfaceCall,
                                               seq_nr=rownr,
                                               status='NEW',
                                               field_01=values[0],
                                               field_02=values[1],
                                               field_03=values[2],
                                               field_04=values[3],
                                               field_05=values[4],
                                               field_06=values[5],
                                               field_07=values[6],
                                               field_08=values[7],
                                               field_09=values[8],
                                               field_10=values[9],
                                               field_11=values[10],
                                               field_12=values[11],
                                               field_13=values[12],
                                               field_14=values[13],
                                               field_15=values[14],
                                               field_16=values[15],
                                               field_17=values[16],
                                               field_18=values[17],
                                               field_19=values[18],
                                               field_20=values[19],
                                               field_21=values[20],
                                               field_22=values[21],
                                               field_23=values[22],
                                               field_24=values[23],
                                               field_25=values[24],
                                               field_26=values[25],
                                               field_27=values[26],
                                               field_28=values[27],
                                               field_29=values[28],
                                               field_30=values[29],
                                               field_31=values[30],
                                               field_32=values[31],
                                               field_33=values[32],
                                               field_34=values[33],
                                               field_35=values[34],
                                               field_36=values[35],
                                               field_37=values[36],
                                               field_38=values[37],
                                               field_39=values[38],
                                               field_40=values[39],
                                               field_41=values[40],
                                               field_42=values[41],
                                               field_43=values[42],
                                               field_44=values[43],
                                               field_45=values[44],
                                               field_46=values[45],
                                               field_47=values[46],
                                               field_48=values[47],
                                               field_49=values[48],
                                               field_50=values[49],
                                               )
    return receivedData


def handle_negometrix_row(row_nr: int,
                          row_values: Tuple[str],
                          interfaceCall: InterfaceCall,
                          field_positions: Dict[str, int],
                          mandatory_field_positions: Tuple[int]):
    logger.debug(f"register ReceivedData {row_nr} - {row_values}")

    receivedData = register_in_recevied_data(row_nr,
                                             row_values,
                                             interfaceCall)
    try:
        status, message = register_contract(row_nr,
                                            row_values,
                                            interfaceCall,
                                            field_positions,
                                            mandatory_field_positions)
    except Exception as ex:
        receivedData.status = ERROR
        receivedData.message = str(ex)
    else:
        receivedData.status = status
        receivedData.message = message
    receivedData.save()


class InterfaceFile(object):

    def __init__(self, file):
        self.file = file

    def process(self):
        # register file in InterfaceCall
        interfaceCall = InterfaceCall.objects.create(filename=self.file.name,
                                                     status=NEW,
                                                     date_time_creation=Now(),
                                                     type=INTERFACE_TYPE_FILE)
        try:
            has_excel_extension, msg = file_has_excel_extension(self.file.name)
            if not has_excel_extension:
                raise Exception(msg)

            is_excel, message = file_is_excel_file(self.file)
            if not is_excel:
                raise Exception(message)

            handle_uploaded_excel_file(self.file, interfaceCall)
            interfaceCall.status = OK
        except Exception as ex:
            interfaceCall.status = ERROR
            interfaceCall.message = f'Reason: {ex.__str__()}'

        interfaceCall.save()

        if not interfaceCall.status == OK:
            raise Exception(interfaceCall.message)


class NegometrixFile(InterfaceFile):
    pass


def create(file):
    return NegometrixFile(file)
